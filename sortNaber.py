import os
import openpyxl

def sort_excel_sheet(file_path):
    # Excelファイルを開く
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Redmine出力ファイル']
    
    # A列のデータとその行全体を取得
    rows = list(sheet.iter_rows(min_row=7, values_only=True))
    header = list(sheet.iter_rows(min_row=6, max_row=6, values_only=True))[0]  # ヘッダー行を取得
    
    # 行のデータをA列の値でソート
    sorted_rows = sorted(rows, key=lambda row: row[0])
    
    # シートの内容をクリア
    for row in sheet.iter_rows(min_row=7, max_col=sheet.max_column, max_row=sheet.max_row):
        for cell in row:
            cell.value = None
    
    # ヘッダーを設定
    for col_num, value in enumerate(header, start=1):
        sheet.cell(row=6, column=col_num, value=value)
    
    # ソートしたデータをシートに書き込む
    for i, row in enumerate(sorted_rows, start=7):
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i, column=j, value=value)
    
    # ファイルを保存して閉じる
    wb.save(file_path)
    wb.close()

def main():
    # スクリプトと同じ階層の「output」フォルダを指定
    folder_path = os.path.join(os.path.dirname(__file__), 'output')
    
    if not os.path.exists(folder_path):
        print("「output」フォルダが見つかりません。")
        return
    
    # 選択したフォルダ内の Excel ファイルを処理
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(folder_path, file_name)
            sort_excel_sheet(file_path)
    
    print("すべてのファイルのデータが昇順に並べ替えられました。")

if __name__ == "__main__":
    main()
