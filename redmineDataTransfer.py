


import os
import openpyxl
import tkinter as tk
from tkinter import filedialog

def get_last_non_empty_cell(ws, rows):
    """指定された行の中で最後に値が入力されているセルを取得"""
    last_cells = {}
    for col in range(1, ws.max_column + 1):
        last_cell = None
        for row in reversed(rows):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None and str(cell.value).strip() != "":
                last_cell = cell
                break
        if last_cell:
            last_cells[col] = last_cell
    return last_cells

def find_header(ws, header, exact_match=False):
    """ヘッダーを見つける（完全一致または部分一致）"""
    for cell in ws[6]:
        if (exact_match and header == str(cell.value)) or (not exact_match and header in str(cell.value)):
            return cell
    return None

def copy_multiple_data_with_mapping(ws_data, ws_target, target_header_rows=[8, 9, 10, 11, 12]):
    # マッピングの定義
    mapping = {
        "ステータス": "ステータス",
        "シナリオID": "シナリオID\n※総合テスト、受入テスト時のみ利用",
        "テストケースNo": "テストケースNo.\n※テストケースを識別できる番号やIDを記載",
        "障害検知日": "障害検知日",
        "起票事業者": "起票事業者",
        "起票者": "起票者",
        "説明": "問題事象・内容\n※詳細な発生要因についても記載すること",
        "優先度": "緊急度",
        "障害起因": "障害起因",
        "障害起因事業者名": "障害起因事業者名\n※他事業者に\n起因する障害\nの場合、選択\n必須",
        "原因切り分けが必要と判断した根拠": "原因切り分けが必要と判断した根拠",
        "原因記入者": "原因記入者",
        "原因工程": "原因工程",
        "原因分類": "原因分類",
        "同件障害No": "同件障害No.\n※調査の結果、「同件障害」と判断された場合はその親となる障害管理No.を記載",
        "不備を抽出すべきテスト工程": "不備を抽出すべきテスト工程",
        "直接原因": "直接原因",
        "根本原因": "根本原因",
        "類似調査要否": "類似調査要否",
        "要否判断の根拠": "要否判断の根拠・類似調査方針",
        "調査結果": "調査結果",
        "類似調査完了日": "完了日",
        "障害対応方針記入者": "障害対応\n方針記入者",
        "障害対応方針": "障害対応方針",
        "変更管理ID or 修正履歴ID": "変更管理ID\nor\n修正履歴ID\n※「設計書修正要否」が要の場合記載必須",
        "暫定対応者": "暫定対応者",
        "暫定対応完了予定日": "暫定対応完了予定日",
        "暫定対応完了日": "暫定対応完了日",
        "暫定対応リソース名": "暫定対応リソース名",
        "暫定対応内容": "暫定対応内容",
        "恒久対応者": "恒久対応者",
        "恒久対応完了予定": "恒久対応完了予定日",
        "恒久対応完了日": "恒久対応完了日",
        "恒久対応リソース名": "恒久対応リソース名",
        "恒久対応内容": "恒久対応内容",
        "リグレッション確認実施要否": "実施要否",
        "リグレッション確認実施内容／対象範囲": "実施内容/対象範囲",
        "リグレッション確認不要と判断した根拠": "不要と判断した根拠",
        "リリース完了予定日": "リリース完了予定日",
        "リリース完了日": "リリース完了日",
        "検証事業者": "検証事業者",
        "検証者": "検証者",
        "検証完了日": "検証完了日",
        "最終確認完了日": "最終確認完了日"
    }

    # 完全一致が必要な項目
    exact_match_headers = {
        "検証完了日",
        "直接原因",
        "障害対応方針",
        "リリース完了予定日",
        "リリース完了日",
        "検証事業者",
        "検証者"
    }

    print("マッピング:")
    print(mapping)

    # Redmine出力ファイルシートのヘッダー行は常に6行目
   
    print("CSVをEY障害管理簿に転記します")
    headers = [cell.value for cell in ws_data[6]]
    print(headers)

    # 【別紙15】障害管理簿シートのヘッダー行を取得
    last_header_cells = get_last_non_empty_cell(ws_target, target_header_rows)
    #print("【別紙15】障害管理簿シートのヘッダー:")
    for col, cell in last_header_cells.items():
        print(f"列 {col}: {cell.value} (行 {cell.row})")

    for header, target_header in mapping.items():
        exact_match = header in exact_match_headers
        header_cell = find_header(ws_data, header, exact_match=exact_match)

        if header_cell:
            print(f"見つけた '{header}' ヘッダー: {header_cell.value}")

            # 最後の行を探す
            last_row = ws_data.max_row
            while ws_data.cell(row=last_row, column=header_cell.column).value is None and last_row > header_cell.row:
                last_row -= 1

            data_cell = None
            for col, last_cell in last_header_cells.items():
                if last_cell.value == target_header:
                    data_cell = last_cell
                    break

            if data_cell:
                # print(f"【別紙15】障害管理簿シートに '{target_header}' ヘッダーを見つけた: {data_cell.value}")

                # データを貼り付ける行を設定
                dest_row = 13

                for row in range(header_cell.row + 1, last_row + 1):
                    cell_value = ws_data.cell(row=row, column=header_cell.column).value
                    if cell_value is not None and str(cell_value).strip() != "":
                        ws_target.cell(row=dest_row, column=data_cell.column).value = cell_value
                        dest_row += 1
            else:
                print(f"【別紙15】障害管理簿シートから '{target_header}' 項目を探せませんでした。")
        else:
            print(f"Redmine出力ファイルシートから '{header}' 項目を探せませんでした。")

    # "他領域への影響有無" 項目処理
    impact_range_cell = find_header(ws_data, "他領域への影響有無", exact_match=False)

    if impact_range_cell:
        last_row = ws_data.max_row
        while ws_data.cell(row=last_row, column=impact_range_cell.column).value is None and last_row > impact_range_cell.row:
            last_row -= 1

        dest_row = 13
        for row in range(impact_range_cell.row + 1, last_row + 1):
            cell_value = ws_data.cell(row=row, column=impact_range_cell.column).value
            if cell_value is not None and str(cell_value).strip() != "":
                ws_target.cell(row=dest_row, column=impact_range_cell.column).value = cell_value
                dest_row += 1

def process_excel_files_in_folder(folder_path):
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".xlsx") or file_name.endswith(".xlsm"):
            file_path = os.path.join(folder_path, file_name)
            try:
                wb = openpyxl.load_workbook(file_path)
                ws_data = wb["Redmine出力ファイル"]
                ws_target = wb["【別紙15】障害管理簿"]

                if ws_data and ws_target:
                    copy_multiple_data_with_mapping(ws_data, ws_target)
                    wb.save(file_path)
            except Exception as e:
                print(f"ファイル {file_name} の処理中にエラーが発生しました: {e}")

def select_folder():
    # フォルダパスを固定する
    return os.path.join(os.path.dirname(__file__), 'output')
if __name__ == "__main__":
    folder_path = select_folder()
    if folder_path:
        process_excel_files_in_folder(folder_path)
        print("RedmineのデータをExcelに転記しました。最後に図形を転記します。しばらくお待ちください。")
    else:
        print("フォルダが選択されませんでした。")
